from flask import  render_template, request, jsonify,redirect,url_for,make_response,send_file,session,Blueprint,current_app
from datetime import datetime
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook,Workbook
from sqlalchemy import desc
from weasyprint import HTML,CSS
from decorators import token_required
import io
import os

from models import *
api_products_bp = Blueprint('api_products', __name__, url_prefix='/api/products')


@api_products_bp.route("",methods=["GET","POST"])
@token_required
def getProducts(user_id):
    products_query = Product.query.filter(Product.user_id==user_id)
    query=request.args.get("search")
    quantity_filter=request.args.get("quantity_filter",type=int)
    if query:
        products_query = products_query.filter(
        (Product.name.like(f"%{query}%")) |
        (Product.barcode.like(f"%{query}%")))

        
    if quantity_filter==1:
        products_query = products_query.filter(Product.quantity_float==0.0)
    elif quantity_filter==2:
        products_query = products_query.filter(Product.quantity_float>0.0)
    
        
    products=products_query.order_by(desc(Product.product_id)).all()
    
    results_list=[
        {"id": p.product_id, "name": p.name, "barcode": p.barcode, "price": p.current_price,"quantity":p.quantity_float} for p in products]
    return jsonify({
              "success": True,
              "results": results_list
     })
@api_products_bp.route("/add", methods=["POST"])
@token_required
def insertemptyproduct(user_id):
    product_price = float(request.json["product_price"])
    product_name=request.json["product_name"]
    product_brcode=request.json["product_brcode"]
    if product_brcode=="":
       product_brcode=datetime.now().strftime("%Y%m%d%H%M%S%f")
    product = Product.query.filter_by(barcode=product_brcode).first()
    
    if product:
        return jsonify({
            "success": True,
            "status": "product exist",
            "product_name": product_name,
            "product_price": product_price,
            "product_brcode":product_brcode
        })
    else:
         
         
         
         new_product=Product(name=product_name,current_price=product_price,barcode=product_brcode,user_id=user_id)
         product = Product.query.filter_by(barcode=product_brcode).first()
         #print(product)
         db2.session.add(new_product)
         db2.session.commit()
         print(new_product.product_id)
         return jsonify({
              "success": True,
              "status": "product added",
              "product_name": product_name,
              "product_price": product_price,
              "product_brcode":product_brcode
          })
@api_products_bp.route("/<int:product_id>/update", methods=["GET", "POST"])
@token_required
def edit_product(user_id,product_id):
    
    product = Product.query.get_or_404(product_id)

    if request.method == "POST":
       if product:
           product.name = request.json["product_name"]
           product.current_price= float(request.json["product_price"])
           product.barcode = request.json["product_brcode"]
           db2.session.commit()
       return jsonify({
            "success": True,
            "status": "product updated",
            "product_name": product.name,
            "product_price": product.current_price,
            "product_brcode":product.barcode
        })

    return jsonify({
            "success": False,
            "status": "Data missing"
        })
@api_products_bp.route("/<int:product_id>/remove", methods=["GET"])
@token_required
def remove_product(user_id,product_id):
    
    product = Product.query.get_or_404(product_id)
    if product:
           db2.session.delete(product)
           db2.session.commit()
           return jsonify({
            "success": True,
            "status": "product deleted"
           })

    return jsonify({
            "success": False,
            "status": "Data missing"
        })
@api_products_bp.route("/search",methods=["POST"])
@token_required
def search_product(user_id):
    target=request.json["field"]
    print(request.json)
    if target=="barcode":
        barcode=request.json["value"]
        product:Product=Product.query.filter(Product.barcode==barcode).first()
    else:
        product_id=int(request.json["value"])  
        product:Product=Product.query.filter(Product.product_id==product_id).first()    
    if product:
        return jsonify({
            "success":True,
            "message":"product found",
            "product_id":product.product_id,
            "product_name":product.name,
            "product_barcode":product.barcode,
            "product_quantity":product.quantity_float
        })
    else:
        return jsonify({
            "success":False,
            "message":"product not found"
        })        
@api_products_bp.route("/export", methods=["GET"])
@token_required
def export_products(user_id):
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Write header row
    ws.append(["Name", "Barcode", "Quantity", "Price", "Purchase Price"])


    #products filter
    products_query = Product.query.filter(Product.user_id==user_id)
    query=request.args.get("search")
    quantity_filter=request.args.get("quantity_filter",type=int)
    if query:
        products_query = products_query.filter(
        (Product.name.like(f"%{query}%")) |
        (Product.barcode.like(f"%{query}%")))

        
    if quantity_filter==1:
        products_query = products_query.filter(Product.quantity_float==0.0)
    elif quantity_filter==2:
        products_query = products_query.filter(Product.quantity_float>0.0)
    
        
    products=products_query.order_by(desc(Product.product_id)).all()

    for p in products:
        # Get latest batch purchase price (or None if no batch exists)
        item = PurchaseItems.query.filter_by(product_id=p.product_id).order_by(PurchaseItems.purchase_id.desc()).first()
        purchase_price = item.purchase_price if item else None

        ws.append([p.name, p.barcode, p.quantity_float, p.current_price, purchase_price])

    # Save to in-memory file
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="products.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@api_products_bp.route("/products.pdf",methods=["GET"])
@token_required
def products_pdf(user_id):
    products_query = Product.query.filter(Product.user_id==user_id)
    query=request.args.get("search")
    quantity_filter=request.args.get("quantity_filter",type=int)
    if query:
        products_query = products_query.filter(
        (Product.name.like(f"%{query}%")) |
        (Product.barcode.like(f"%{query}%")))

        
    if quantity_filter==1:
        products_query = products_query.filter(Product.quantity_float==0.0)
    elif quantity_filter==2:
        products_query = products_query.filter(Product.quantity_float>0.0)
    
        
    products=products_query.order_by(desc(Product.product_id)).all()

    # render to PDF
    html = render_template("products_list_pdf_template.html", data=products)
    
    css_path = os.path.join(current_app.root_path, "static", "css","bootstrap.min.css")
    
    pdf = HTML(string=html,base_url=current_app.root_path).write_pdf()
    #pdf_bytes = HTML(string=html, base_url=app.root_path).write_pdf()
    pdf_bytes = HTML(string=html, base_url=current_app.root_path).write_pdf(stylesheets=[CSS(css_path)])
    pdf_file = io.BytesIO(pdf_bytes)
    return send_file(
        pdf_file,
        mimetype="application/pdf",
        as_attachment=True,          # True → download, False → open in browser
        download_name="products.pdf"  # filename for the browser
    )	
@api_products_bp.route("/import", methods=["POST"])
@token_required
def upload_products(user_id):
    try:
    	
        file=5
        if "file" not in request.files:
            return jsonify({
            "success": False,
            "status": "File Upload Failed"
            })

        file = request.files["file"]

    except Exception as e:
        print(e)
    # Load the Excel file
    wb = load_workbook(file)
    ws = wb.active   # first sheet
    name="Opening_"+datetime.now().strftime("%Y%m%d_%H%M")
    
    supplier=Suppliers(name=name,email="",phone="",user_id=user_id)
    db2.session.add(supplier)
    db2.session.flush()
    purchase=Purchases(supplier_id=supplier.supplier_id,user_id=user_id)
    db2.session.add(purchase)
    db2.session.flush()
    # Skip headers (row 1)
    total=0
    for row in ws.iter_rows(min_row=2, values_only=True):
          print("  ..... ,  ",row)
          name,barcode,quantity,price,purchase_price=row
          quantity=float(quantity)
          purchase_price=float(purchase_price)
          price=float(price)
          pd=Product.query.filter(Product.barcode==barcode).first()
          if pd:
                print("barcode exists")
                pd.quantity_float += quantity
                pd.current_price = price  # if you want latest price
          else:
               pd=Product(name=name,barcode=barcode,quantity_float=quantity,current_price=price,user_id=user_id)
               db2.session.add(pd)
               db2.session.flush()
          total+=quantity*purchase_price
          new_prod_batch=ProductBatches(product_id=pd.product_id,purchase_price=purchase_price,quantity_float=quantity)
          db2.session.add(new_prod_batch)
              
          
          new_purchase_item=PurchaseItems(product_id=pd.product_id,purchase_price=purchase_price,purchase_id=purchase.purchase_id,quantity_float=quantity,remain_quantity=quantity)
          db2.session.add(new_purchase_item)
              
    purchase.total_amount=total   
    db2.session.commit()
        
        
    
    return jsonify({
            "success": True,
            "status": "Products Uploaded"
     })
    
    