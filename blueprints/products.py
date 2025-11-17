from flask import  render_template, request, jsonify,redirect,url_for,make_response,send_file,session,Blueprint
from datetime import datetime
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook,Workbook
from sqlalchemy import desc
from weasyprint import HTML,CSS
import io
import os

from models import *
products_bp = Blueprint('products', __name__, url_prefix='/products')



@products_bp.route("/scanner",methods=["GET"])
def scanner():
    if "user" not in session: return redirect(url_for("users.login"))
    return render_template("scanbr.html")
 
@products_bp.route("/manageproducts",methods=["GET"])
def manageproducts():
    if "user" not in session: return redirect(url_for("users.login"))
    return render_template("manageproducts.html")

@products_bp.route("/product/<int:product_id>/addproducts",methods=["GET", "POST"])
def addproducts(product_id):
    if "user" not in session: return redirect(url_for("users.login"))
    product=None
    print(product_id)
    if product_id>0:
    	
          product = Product.query.get_or_404(product_id)
    return render_template("addproduct.html",product=product)

@products_bp.route("/insertproductpurchase", methods=["POST"])
def insertproductpurchase():
    if "user" not in session: return redirect(url_for("users.login"))
    product_price = float(request.json["product_price"])
    product_purchase_price = float(request.json["product_purchase_price"])
    
    product_quantity= float(request.json["product_quantity"])
    
    product_name=request.json["product_name"]
    product_brcode=request.json["product_brcode"]
    purchase_id=request.json["purchase_id"]
    purchase=Purchases.query.filter(Purchases.purchase_id==purchase_id).first()
    if product_brcode=="":
       product_brcode=datetime.now().strftime("%Y%m%d%H%M%S%f")
    product = Product.query.filter_by(barcode=product_brcode).first()
    if product:
        if purchase and purchase.status:
              product.quantity_float=product.quantity_float+product_quantity
        product.current_price=product_price
        db2.session.commit()
        ##batch=ProductBatches(product_id=product.product_id,purchase_price=product_purchase_price,quantity_float=product_quantity)
        ##db2.session.add(batch)
        ##db2.session.commit()
        
        if  purchase:
               purchase_item=PurchaseItems(purchase_id=purchase_id,product_id=product.product_id,purchase_price=product_purchase_price,quantity_float=product_quantity,remain_quantity=product_quantity)
               db2.session.add(purchase_item)
               purchase.total_amount+=product_purchase_price*product_quantity
        db2.session.commit()
        return jsonify({
            "success": True,
            "status": "product updated",
            "product_name": product_name,
            "product_price": product_price,
            "product_brcode":product_brcode
        })
    else:
         
         
         
         new_product=Product(name=product_name,current_price=product_price,barcode=product_brcode,user_id=int(session["user_id"]))
         product = Product.query.filter_by(barcode=product_brcode).first()
         #print(product)
         db2.session.add(new_product)
         db2.session.commit()
         print(new_product.product_id)
         product = Product.query.filter_by(barcode=product_brcode).first()
         #print(product,product.product_id,product.name)        
         ##batch=ProductBatches(product_id=product.product_id,purchase_price=product_purchase_price,quantity_float=product_quantity)
         ##db2.session.add(batch)
         ##db2.session.commit()
         #print(product,product.product_id)
         #purchase=Purchases.query.filter(Purchases.purchase_id==purchase_id).first()
         if  purchase:
               purchase_item=PurchaseItems(purchase_id=purchase_id,product_id=product.product_id,purchase_price=product_purchase_price,quantity_float=product_quantity,remain_quantity=product_quantity)
               db2.session.add(purchase_item)
               purchase.total_amount+=product_purchase_price*product_quantity
         db2.session.commit()
         
         
         return jsonify({
              "success": True,
              "status": "product added",
              "product_name": product_name,
              "product_price": product_price,
              "product_brcode":product_brcode
          })
    
@products_bp.route("/insertemptyproduct", methods=["POST"])
def insertemptyproduct():
    if "user" not in session: return redirect(url_for("users.login"))
    product_price = float(request.json["product_price"])
    product_name=request.json["product_name"]
    product_brcode=request.json["product_brcode"]
    if product_brcode=="":
       product_brcode=datetime.now().strftime("%Y%m%d%H%M%S%f")
    product = Product.query.filter_by(barcode=product_brcode).first()
    
    if product:
        return jsonify({
            "success": True,
            "status": "product updated",
            "product_name": product_name,
            "product_price": product_price,
            "product_brcode":product_brcode
        })
    else:
         
         
         
         new_product=Product(name=product_name,current_price=product_price,barcode=product_brcode,user_id=int(session["user_id"]))
         product = Product.query.filter_by(barcode=product_brcode).first()
         #print(product)
         db2.session.add(new_product)
         db2.session.commit()
         print(new_product.product_id)
         return jsonify({
              "success": True,
              "status": "product added",
              "product_name": product_name,
              "product_price": product_price,
              "product_brcode":product_brcode
          })
    


@products_bp.route("/productlist",methods=["GET","POST"])
def productlist():
    if "user" not in session: return redirect(url_for("users.login"))
    products_query = Product.query.filter(Product.user_id==int(session["user_id"]))
    
    if request.method=="POST":
        query = request.json["search_q"]
        products_filter = int(request.json["products_filter"])
        if query:
            products_query = products_query.filter(
            (Product.name.like(f"%{query}%")) |
            (Product.barcode.like(f"%{query}%")))
        
        if products_filter==1:
           products_query = products_query.filter(Product.quantity_float==0.0)
        elif products_filter==2:
        	products_query = products_query.filter(Product.quantity_float>0.0)
        products=products_query.order_by(desc(Product.product_id)).all()
        results_list=[
        {"id": p.product_id, "name": p.name, "barcode": p.barcode, "price": p.current_price,"quantity":p.quantity_float} for p in products]
        return jsonify({
              "success": True,
              "results": results_list
        })
    products=products_query.order_by(desc(Product.product_id)).all()
    response = make_response(render_template("products_list.html",data=products,mode="manage"))
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response
@products_bp.route("/search")
def search():
    if "user" not in session: return redirect(url_for("users.login"))
    query = request.args.get("q", "")

    results = Product.query.filter(
        ((Product.name.like(f"%{query}%")) |
        (Product.barcode.like(f"%{query}%"))) &(Product.user_id==int(session["user_id"]))
    ).all()
    print(results)
    return render_template("search_product.html", results=results, query=query)
@products_bp.route("/search2",methods=["POST"])
def search2():
    if "user" not in session: return redirect(url_for("users.login"))
    query = request.json["query"]
    sale_id=int(request.json["sale_id"])
    results = Product.query.filter(
        ((Product.name.like(f"%{query}%")) |
        (Product.barcode.like(f"%{query}%"))) &(Product.user_id==int(session["user_id"]))
    ).all()
    print(results)
    if len(results)==0:
       return jsonify({
              "success": False,
              "results": [],
              "total":0
       })
    results_list=[
        {"id": p.product_id, "name": p.name, "barcode": p.barcode, "price": p.current_price}
        for p in results
    ]
    rs=SaleItems.query.filter(SaleItems.sale_id==sale_id,SaleItems.product_id==results[0].product_id).first()
    if rs:
    	rs.quantity_float+=1
    else:
        new_sale_item=SaleItems(sale_id=sale_id,product_id=results[0].product_id,unit_price=results[0].current_price,quantity_float=1)
        db2.session.add(new_sale_item)
    db2.session.commit()
    
    results2 = SaleItems.query.filter(SaleItems.sale_id==sale_id).all()
    result_list2=[]
    for r in results2:
       print(r.product_id)
       if r.product_id:
       	p = Product.query.filter(Product.product_id==r.product_id,Product.user_id==int(session["user_id"])).first()
       	x={"item_id": r.item_id, "name": p.name, "barcode": p.barcode, "price": r.unit_price,"quantity":r.quantity_float}
       else:
           x={"item_id": r.item_id, "name": "", "barcode": "",  "price": r.unit_price,"quantity":r.quantity_float}    
           
       result_list2.append(x)
    print(len(results2))
    total=0
    for st in results2:
    	total+=st.unit_price*st.quantity_float
    print(total)
    
    return jsonify({
              "success": True,
              "results": result_list2,
              "total":total
    })
@products_bp.route("/product/<int:product_id>/edit", methods=["GET", "POST"])
def edit_product(product_id):
    if "user" not in session: return redirect(url_for("users.login"))
    product = Product.query.get_or_404(product_id)

    if request.method == "POST":
        product.name = request.json["product_name"]
        product.current_price= float(request.json["product_price"])
        product.barcode = request.json["product_brcode"]

        db2.session.commit()

        return jsonify({
            "success": True,
            "status": "product updated",
            "product_name": product.name,
            "product_price": product.current_price,
            "product_brcode":product.barcode
        })

    return render_template("modifyproduct.html", product=product)
@products_bp.route("/product/<int:product_id>/remove", methods=["GET", "POST"])
def remove_product(product_id):
    if "user" not in session: return redirect(url_for("users.login"))
    product = Product.query.get(product_id)
    
    if request.method == "POST":
        if product:
            db2.session.delete(product)
            db2.session.commit()
        return """
    <script>
        // Close current window
        // Optional: redirect parent window if opened as popup
        if (window.opener) {
             window.opener.location.reload() ;
        }
        window.close();
    </script>
    """
    return render_template("confirm_delete.html", product=product)
@products_bp.route("/product/<int:product_id>/viewpurchases", methods=["GET", "POST"])
def view_purchases(product_id):
    if "user" not in session: return redirect(url_for("users.login"))
    db = get_db(); cursor = db.cursor()
    cursor.execute("""
SELECT p.product_id, p.barcode,
p.name,p.current_price,
it.purchase_price, it.quantity_float,pu.purchase_date 
FROM purchases pu
JOIN purchase_items it
ON pu.purchase_id = it.purchase_id 
JOIN products p
ON it.product_id = p.product_id 
where it.product_id=%s 
order by pu.purchase_date DESC
""",(product_id,))
    results=cursor.fetchall()
    print(len(results))
    return render_template("product_purchases.html",data=results)
@products_bp.route("/product/<int:product_id>/viewsales", methods=["GET", "POST"])
def view_product_sales(product_id):
    if "user" not in session: return redirect(url_for("users.login"))
    db = get_db(); cursor = db.cursor()
    cursor.execute("""
SELECT p.product_id, p.barcode,
p.name,p.current_price,
it.unit_price, it.quantity_float,sa.sale_date 
FROM sales sa
JOIN sale_items it
ON sa.sale_id = it.sale_id 
JOIN products p
ON it.product_id = p.product_id 
where it.product_id=%s 
order by it.item_id DESC
""",(product_id,))
    results=cursor.fetchall()
    print(len(results))
    return render_template("product_sales.html",data=results)
@products_bp.route("/product/select/<string:type>/<int:row_id>", methods=["GET", "POST"])
def selectproduct(type,row_id):
    if "user" not in session: return redirect(url_for("users.login"))
    products = Product.query.filter(Product.user_id==int(session["user_id"])).order_by(desc(Product.product_id)).all()
    response = make_response(render_template("products_list.html",data=products,mode="select",type=type,row_id=row_id))
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response
@products_bp.route("/product/selected", methods=["GET", "POST"])
def selectedproduct():
    if "user" not in session: return redirect(url_for("users.login"))
    product_id=int(request.json["product_id"])
    row_id=int(request.json["row_id"])
    type=request.json["type"]	
    print("       type:",type)
    print("       row_id:",row_id)
    print("       product_id:",product_id)
    product = Product.query.filter(Product.product_id==product_id).first()
    if product:
        if "sale_items" in type:
            item=SaleItems.query.filter(SaleItems.item_id==row_id).first()
            if item:
               item.product_id=product.product_id
               item.description=product.name
               print("            product.quantity_float:",product.quantity_float)
               print("            item.quantity_float:",item.quantity_float)
               if type!="sale_items_before":
                    product.quantity_float-=item.quantity_float
               print("            product.quantity_float:",product.quantity_float)
               purchase_item=PurchaseItems.query.filter(PurchaseItems.product_id==item.product_id).order_by(desc(PurchaseItems.purchase_item_id)).first()
               if purchase_item:
                  item.profit=item.unit_price-purchase_item.purchase_price
        elif type=="purchase_items":
            purchase=Purchases.query.filter(Purchases.purchase_id==row_id,Purchases.user_id==int(session["user_id"])).first()
            if purchase:
            	print("&&&&&&&&&&&")
            	x={"product_id": product.product_id, "name": product.name, "barcode": product.barcode, "price": product.current_price}
            	return jsonify({
                 "success": True,
                 "results":x
                })			
        db2.session.commit()
    return jsonify({
              "success": True
    })		

@products_bp.route("/product/export", methods=["GET"])
def export_products():
    if "user" not in session: return redirect(url_for("users.login"))
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Write header row
    ws.append(["Name", "Barcode", "Quantity", "Price", "Purchase Price"])

    # Fetch all products
    products = Product.query.filter(Product.user_id==int(session["user_id"])).all()

    for p in products:
        # Get latest batch purchase price (or None if no batch exists)
        item = PurchaseItems.query.filter_by(product_id=p.product_id).order_by(PurchaseItems.purchase_id.desc()).first()
        purchase_price = item.purchase_price if item else None

        ws.append([p.name, p.barcode, p.quantity_float, p.current_price, purchase_price])

    # Save to in-memory file
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="products.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@products_bp.route("/products.pdf/<string:filt>/", defaults={"query": ""})
@products_bp.route("/products.pdf/<string:filt>/<string:query>/")
def products_pdf(filt,query=""):
    if "user" not in session: return redirect(url_for("users.login"))
    # 🔹 directly reuse the same query
    products = Product.query.filter(Product.user_id==int(session["user_id"])).order_by(desc(Product.product_id)).all()
    if query:
        products = Product.query.filter(
          ((Product.name.like(f"%{query}%")) |
          (Product.barcode.like(f"%{query}%"))) & (Product.user_id==int(session["user_id"]))
          ).order_by(desc(Product.product_id)).all()
    products_filte=int(filt)      
    if products_filte==1:
           products = Product.query.filter(Product.quantity_float==0.0,Product.user_id==int(session["user_id"])).order_by(desc(Product.product_id)).all()
    elif products_filte==2:
           products = Product.query.filter(Product.quantity_float>0.0,Product.user_id==int(session["user_id"])).order_by(desc(Product.product_id)).all()
    
 
    # render to PDF
    html = render_template("products_list_pdf_template.html", data=products)
    
    css_path = os.path.join(app.root_path, "static", "css","bootstrap.min.css")
    
    pdf = HTML(string=html,base_url=app.root_path).write_pdf()
    #pdf_bytes = HTML(string=html, base_url=app.root_path).write_pdf()
    pdf_bytes = HTML(string=html, base_url=app.root_path).write_pdf(stylesheets=[CSS(css_path)])
    pdf_file = io.BytesIO(pdf_bytes)
    return send_file(
        pdf_file,
        mimetype="application/pdf",
        as_attachment=True,          # True → download, False → open in browser
        download_name="products.pdf"  # filename for the browser
    )	
    
@products_bp.route("/product/import", methods=["GET", "POST"])
def upload_products():
    if "user" not in session: return redirect(url_for("users.login"))
    if request.method == "POST":
        file = request.files["file"]
        if not file:
            return "No file uploaded", 400
        
    
        # Load the Excel file
        wb = load_workbook(file)
        ws = wb.active   # first sheet
        name="Opening_"+datetime.now().strftime("%Y%m%d_%H%M")
        
        supplier=Suppliers(name=name,email="",phone="",user_id=int(session["user_id"]))
        db2.session.add(supplier)
        db2.session.flush()
        purchase=Purchases(supplier_id=supplier.supplier_id,user_id=int(session["user_id"]))
        db2.session.add(purchase)
        db2.session.flush()
        # Skip headers (row 1)
        total=0
        for row in ws.iter_rows(min_row=2, values_only=True):
              print("  ..... ,  ",row)
              name,barcode,quantity,price,purchase_price=row
              quantity=float(quantity)
              purchase_price=float(purchase_price)
              price=float(price)
              pd=Product.query.filter(Product.barcode==barcode).first()
              if pd:
                    print("barcode exists")
                    pd.quantity_float += quantity
                    pd.current_price = price  # if you want latest price
              else:
                   pd=Product(name=name,barcode=barcode,quantity_float=quantity,current_price=price,user_id=int(session["user_id"]))
                   db2.session.add(pd)
                   db2.session.flush()
              total+=quantity*purchase_price
              new_prod_batch=ProductBatches(product_id=pd.product_id,purchase_price=purchase_price,quantity_float=quantity)
              db2.session.add(new_prod_batch)
                  
              
              new_purchase_item=PurchaseItems(product_id=pd.product_id,purchase_price=purchase_price,purchase_id=purchase.purchase_id,quantity_float=quantity,remain_quantity=product_quantity)
              db2.session.add(new_purchase_item)
                  
        purchase.total_amount=total   
        db2.session.commit()
        return """
    <script>
        history.go(-2)
    </script>
    """

    return render_template("upload.html")
